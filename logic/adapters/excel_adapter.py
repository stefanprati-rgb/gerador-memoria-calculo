"""
Adaptadores para leitura da base Balanço Energético e escrita no template mc.xlsx.
Suporta detecção dinâmica de header e formatação diferenciada para Fatura Pai.

Otimizações de performance:
- Detecção de header com openpyxl read_only (leve)
- Leitura seletiva de colunas (usecols) — ~15 de 125
- Compatível com @st.cache_data no app.py
"""
import pandas as pd
from typing import List, Dict, Any, Optional
from logic.core.mapping import (
    get_base_columns,
    get_required_columns,
    HEADER_MARKER_COLUMNS,
    HEADER_SCAN_ROWS,
    CLIENT_COLUMN,
    PERIOD_COLUMN,
    PARENT_ROW_FLAG,
    OPTIONAL_BASE_COLUMNS,
    ENRICHMENT_KEY,
    ACCOUNT_NUMBER_COL,
    SEPARATOR_ROW_FLAG,
    CLASSIFICATION_COL,
    CHILD_ROW_FLAG,
    CLASSIFICATION_LABEL_REGRA,
)
from logic.core.dates import format_reference_period, format_full_date

import logging

logger = logging.getLogger(__name__)


class ColumnValidationError(Exception):
    """Erro levantado quando colunas obrigatórias não são encontradas na planilha base."""
    pass


class HeaderNotFoundError(Exception):
    """Erro levantado quando o cabeçalho não pode ser detectado automaticamente."""
    pass


def _detect_header_openpyxl(file_path: str, sheet_name: str) -> int:
    """
    Detecta a linha de cabeçalho usando openpyxl em modo read_only (muito leve).
    Retorna o índice 0-based para uso no pandas (header=N).
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=HEADER_SCAN_ROWS, values_only=True)):
        vals = set(str(v).strip() for v in row if v is not None)
        if all(marker in vals for marker in HEADER_MARKER_COLUMNS):
            header_row = row_idx  # 0-based
            break
    
    wb.close()
    
    if header_row is None:
        raise HeaderNotFoundError(
            f"Não foi possível detectar o cabeçalho nas primeiras {HEADER_SCAN_ROWS} linhas. "
            f"Colunas-marcador esperadas: {HEADER_MARKER_COLUMNS}"
        )
    
    return header_row


class BaseExcelReader:
    """Adaptador para leitura da planilha Balanço Energético com detecção dinâmica de header."""

    def __init__(self, file_path_or_buffer: Any, sheet_name: str = "Balanco Operacional"):
        """
        Inicializa o leitor com detecção dinâmica do header e leitura seletiva de colunas.
        
        Otimizações aplicadas:
        1. Header detection via openpyxl read_only (leve, sem carregar dados)
        2. usecols com apenas as colunas necessárias (~15 de 125)
        """
        self.sheet_name = sheet_name

        # Verifica se é um arquivo Parquet de cache
        if isinstance(file_path_or_buffer, str) and file_path_or_buffer.endswith(".parquet"):
            logger.info("Carregando base do cache ultrarrápido Parquet: %s", file_path_or_buffer)
            self.df = pd.read_parquet(file_path_or_buffer, engine="fastparquet")
            self._normalize_columns()
            self._validate_columns()
            logger.info("Base Parquet carregada com %d registros e %d colunas.", len(self.df), len(self.df.columns))
            return
        
        # Detectar header — para file path usa openpyxl (rápido), para buffer usa pandas
        if isinstance(file_path_or_buffer, str):
            header_row = _detect_header_openpyxl(file_path_or_buffer, sheet_name)
        else:
            header_row = self._detect_header_pandas(file_path_or_buffer)
        
        logger.info("Header detectado na linha %d (0-indexed / linha %d no Excel).", header_row, header_row + 1)
        
        # Leitura seletiva — apenas as colunas necessárias
        required_cols = get_required_columns()
        
        try:
            self.df = pd.read_excel(
                file_path_or_buffer,
                sheet_name=sheet_name,
                header=header_row,
                usecols=lambda col: col.strip() in required_cols,
            )
        except Exception:
            # Fallback: se usecols falhar (ex: coluna renomeada), ler tudo
            logger.warning("Leitura seletiva falhou, carregando todas as colunas.")
            self.df = pd.read_excel(file_path_or_buffer, sheet_name=sheet_name, header=header_row)
        
        self._normalize_columns()
        self._validate_columns()
        logger.info("Base carregada com %d registros e %d colunas.", len(self.df), len(self.df.columns))

    def _detect_header_pandas(self, file_buffer: Any) -> int:
        """Fallback para detecção via pandas (para UploadedFile/buffers)."""
        raw = pd.read_excel(
            file_buffer,
            sheet_name=self.sheet_name,
            nrows=HEADER_SCAN_ROWS,
            header=None,
        )
        
        for i, row in raw.iterrows():
            vals = set(str(v).strip() for v in row.values if pd.notna(v))
            if all(marker in vals for marker in HEADER_MARKER_COLUMNS):
                return i
        
        raise HeaderNotFoundError(
            f"Não foi possível detectar o cabeçalho nas primeiras {HEADER_SCAN_ROWS} linhas. "
            f"Colunas-marcador esperadas: {HEADER_MARKER_COLUMNS}"
        )

    def _normalize_columns(self):
        """Remove espaços extras dos nomes de colunas para evitar falhas por diferenças mínimas."""
        self.df.columns = self.df.columns.str.strip()

    @staticmethod
    def _normalize_period_value(value: Any) -> str:
        """Normaliza referência para MM/YYYY; retorna string vazia se inválida."""
        normalized = format_reference_period(value, default="")
        return normalized.strip() if isinstance(normalized, str) else ""

    def _validate_columns(self):
        """Valida se todas as colunas esperadas pelo mapeamento estão presentes na base."""
        expected = get_base_columns()
        missing = [c for c in expected if c not in self.df.columns and c not in OPTIONAL_BASE_COLUMNS]
        if missing:
            raise ColumnValidationError(
                f"Colunas obrigatórias ausentes na planilha base: {missing}. "
                f"Colunas encontradas: {list(self.df.columns)}"
            )

    def get_clients(self) -> List[str]:
        """Retorna lista de clientes (Razao Social) únicos na base, ordenados."""
        if CLIENT_COLUMN not in self.df.columns:
            return []
        clients = self.df[CLIENT_COLUMN].dropna().unique().tolist()
        return sorted([str(c) for c in clients])

    def get_periods(self) -> List[str]:
        """Retorna lista de períodos (Referencia) únicos."""
        if PERIOD_COLUMN not in self.df.columns:
            return []
        normalized_periods = (
            self.df[PERIOD_COLUMN]
            .dropna()
            .map(self._normalize_period_value)
        )

        def _period_sort_key(period: str) -> tuple[int, int, str]:
            # Ordenação cronológica real (ano -> mês), com fallback lexical.
            # Aceita MM/YYYY ou MM-YYYY.
            try:
                normalized = str(period).strip().replace("-", "/")
                month_str, year_str = normalized.split("/")
                month = int(month_str)
                year = int(year_str)
                if 1 <= month <= 12:
                    return (year, month, str(period))
            except Exception:
                pass
            return (9999, 99, str(period))

        valid_periods = sorted({p for p in normalized_periods if p}, key=_period_sort_key)
        return valid_periods

    def filter_data(self, clients: List[str], periods: List[str]) -> pd.DataFrame:
        """Filtra o DataFrame pelos clientes e períodos especificados."""
        mask = pd.Series(True, index=self.df.index)

        if clients:
            client_mask = self.df[CLIENT_COLUMN].isin(clients)

            # Quando há variação de Razão Social para o mesmo documento,
            # inclui todas as linhas do mesmo CPF/CNPJ dos clientes selecionados.
            if "CPF/CNPJ" in self.df.columns:
                def _normalize_doc(v: Any) -> str:
                    if pd.isna(v):
                        return ""
                    return "".join(ch for ch in str(v) if ch.isdigit())

                selected_docs = (
                    self.df.loc[client_mask, "CPF/CNPJ"]
                    .map(_normalize_doc)
                    .replace("", pd.NA)
                    .dropna()
                    .unique()
                    .tolist()
                )
                if selected_docs:
                    docs_mask = self.df["CPF/CNPJ"].map(_normalize_doc).isin(selected_docs)
                    client_mask = client_mask | docs_mask

            mask = mask & client_mask

        if periods:
            selected_periods = set()
            for period in periods:
                normalized = self._normalize_period_value(period)
                if normalized:
                    selected_periods.add(normalized)
            if not selected_periods:
                mask = mask & pd.Series(False, index=self.df.index)
            else:
                period_series = self.df[PERIOD_COLUMN].map(self._normalize_period_value)
                mask = mask & period_series.isin(selected_periods)

        filtered = self.df[mask].copy()
        logger.info("Filtro aplicado: %d clientes, %d períodos → %d registros.", len(clients), len(periods), len(filtered))
        return filtered


class TemplateExcelWriter:
    """Adaptador para escrever dados no template mc.xlsx com formatação de dados."""

    # Colunas da base que contêm valores monetários (R$)
    CURRENCY_COLUMNS = {
        "Valor Enviado Emissão", "Tarifa Raizen", "Custo c/ GD", "Custo s/ GD", "Ganho total Padrão",
    }

    # Colunas da base que contêm datas (mês/ano)
    DATE_COLUMNS = {"Referencia"}

    # Colunas da base que contêm datas completas (dia/mês/ano)
    FULL_DATE_COLUMNS = {"Vencimento", "Data de Pagamento"}

    # Colunas da base que contêm CPF/CNPJ
    DOCUMENT_COLUMNS = {"CPF/CNPJ"}

    # Colunas que devem ser tratadas como texto puro
    TEXT_COLUMNS = {ACCOUNT_NUMBER_COL, ENRICHMENT_KEY, "Referencia"}

    # Mapeamento reverso para localizar colunas no template físico (mc.xlsx) que ainda usam nomes antigos
    LEGACY_HEADER_MAP = {
        "Data  Ref": "Referencia",
        "UC": "No. UC",
        "CNPJ": "CPF/CNPJ",
        "Razão Social": "Razao Social",
        "Energia compensada pela Raízen (kWh)": "Cred. Consumido Raizen",
        "Regra aplicada": "Desconto Contratado",
        "Status financeiro": "Status Pos-Faturamento",
        "Boleto faturado (R$)": "Valor Enviado Emissão",
        "Tarifa Distribuidora": "Tarifa Raizen",
        "Tafira Distribuidora": "Tarifa Raizen",
        "Custo com GD R$": "Custo c/ GD",
        "Custo sem GD R$": "Custo s/ GD",
        "Economia (R$)": "Ganho total Padrão",
        "Nº Conta": ACCOUNT_NUMBER_COL,
        "Número da conta": ACCOUNT_NUMBER_COL,
    }

    def __init__(self, template_path_or_buffer: Any):
        self.template_source = template_path_or_buffer

    @staticmethod
    def _format_date(val) -> Optional[str]:
        """
        Converte datetime/Timestamp ou string MM-YYYY/MM/YYYY em string MM/YYYY.
        Garante que não haja injeção de dias (viagem no tempo).
        """
        if val is None or pd.isna(val):
            return None
        return format_reference_period(val, default=str(val))

    @staticmethod
    def _format_date_full(val) -> Optional[str]:
        """Converte datetime/Timestamp/str em string DD-MM-YYYY ou texto literal."""
        return format_full_date(val)

    @staticmethod
    def _format_document(val) -> Optional[str]:
        """Formata CPF (XXX.XXX.XXX-XX) ou CNPJ (XX.XXX.XXX/XXXX-XX)."""
        if val is None:
            return None
        raw = str(int(val)) if isinstance(val, (int, float)) else str(val)
        raw = raw.strip().replace(".", "").replace("-", "").replace("/", "")

        if len(raw) == 14:
            return f"{raw[:2]}.{raw[2:5]}.{raw[5:8]}/{raw[8:12]}-{raw[12:14]}"
        elif len(raw) == 11:
            return f"{raw[:3]}.{raw[3:6]}.{raw[6:9]}-{raw[9:11]}"
        else:
            # Tentar preencher com zeros à esquerda para CNPJ
            if len(raw) < 14:
                raw = raw.zfill(14)
                return f"{raw[:2]}.{raw[2:5]}.{raw[5:8]}/{raw[8:12]}-{raw[12:14]}"
            return str(val)

    def generate_bytes(self, data_to_insert: pd.DataFrame, column_mapping: Dict[str, str], tipo_apresentacao: str = "Tabela Única", incluir_resumo: bool = False, separar_auditoria: bool = False) -> bytes:
        """
        Lê o template, insere as linhas filtradas e retorna os bytes do Excel gerado.
        Aplica formatação:
        - Datas → MM/YYYY
        - CPF/CNPJ → XX.XXX.XXX/XXXX-XX
        - Valores monetários → R$ #.##0,00
        - Fatura Pai → negrito + fundo amarelo
        """
        import io
        import openpyxl
        from copy import copy

        # Carregar o template
        wb = openpyxl.load_workbook(self.template_source)
        ws = wb.active

        template_ws = ws
        template_ws.title = "_template_temp"

        # 1. Mapear e Enforcar Layout (Strict layout)
        header_row_idx = 1
        original_max_col = template_ws.max_column
        
        for c in range(1, original_max_col + 1):
            template_ws.cell(row=header_row_idx, column=c).value = None

        template_col_to_idx = {}
        for idx, (logical_name, target_label) in enumerate(column_mapping.items(), 1):
            target_label = target_label.strip()
            template_ws.cell(row=header_row_idx, column=idx, value=target_label)
            template_col_to_idx[logical_name] = idx
            
        expected_cols = len(column_mapping)
        if original_max_col > expected_cols:
            template_ws.delete_cols(expected_cols + 1, original_max_col - expected_cols + 50)

        # Preparar dados: Separar Auditoria se ativo
        df_financeiro = data_to_insert
        df_auditoria = pd.DataFrame()
        
        if separar_auditoria:
            mask_aud = data_to_insert.get(CHILD_ROW_FLAG, pd.Series(False, index=data_to_insert.index)).astype(bool) | (data_to_insert.get(CLASSIFICATION_COL, "") == CLASSIFICATION_LABEL_REGRA)
            df_auditoria = data_to_insert[mask_aud].copy()
            df_financeiro = data_to_insert[~mask_aud].copy()

        # Definir grupos de dados
        df_groups = []
        if tipo_apresentacao == "Tabela Única":
            df_groups.append(("Fin - Consolidado" if separar_auditoria else "Consolidado", df_financeiro))
            if separar_auditoria and not df_auditoria.empty:
                df_groups.append(("Aud - Consolidado", df_auditoria))
        else:
            if "Distribuidora" in data_to_insert.columns:
                for name, group in df_financeiro.groupby("Distribuidora", sort=False):
                    prefix = "Fin - " if separar_auditoria else ""
                    df_groups.append((f"{prefix}{name}", group))
                
                if separar_auditoria and not df_auditoria.empty:
                    for name, group in df_auditoria.groupby("Distribuidora", sort=False):
                        df_groups.append((f"Aud - {name}", group))
            else:
                df_groups.append(("Fin - Faturas" if separar_auditoria else "Faturas", df_financeiro))
                if separar_auditoria and not df_auditoria.empty:
                    df_groups.append(("Aud - Faturas", df_auditoria))

        start_row = 2

        # Estilos para a linha "Fatura Pai" — fundo amarelo visível
        parent_font = openpyxl.styles.Font(bold=True, size=11)
        parent_fill = openpyxl.styles.PatternFill(
            start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"
        )

        # Formato monetário brasileiro
        currency_format = '#,##0.00'

        missing_font = openpyxl.styles.Font(
            color="BF360C", bold=True
        )

        total_rows_written = 0
        for name, group_df in df_groups:
            ws = wb.copy_worksheet(template_ws)
            safe_title = "".join([c for c in name if c not in r"\/?*[]:"])[:31]
            ws.title = safe_title if safe_title else "Consolidado"
            current_row = start_row

            for _, row in group_df.iterrows():
                is_parent = bool(row.get(PARENT_ROW_FLAG, False))
                is_separator = bool(row.get(SEPARATOR_ROW_FLAG, False))

                for base_col, col_idx in template_col_to_idx.items():
                    val = None

                    if not is_separator and base_col in row:
                        val = row[base_col]

                        # Tratar NaNs
                        if pd.notna(val):
                            # Tarifa Raizen não pode ser negativa (requisição do usuário)
                            if base_col == "Tarifa Raizen" and isinstance(val, (int, float)) and val < 0:
                                val = 0.0

                            # Aplicar formatação por tipo de coluna
                            if base_col in self.DATE_COLUMNS:
                                val = self._format_date(val)
                            elif base_col in self.FULL_DATE_COLUMNS:
                                val = self._format_date_full(val)
                            elif base_col in self.DOCUMENT_COLUMNS:
                                val = self._format_document(val)
                        else:
                            val = None

                    new_cell = ws.cell(row=current_row, column=col_idx, value=val)

                    # Formato numérico para moeda
                    if base_col in self.CURRENCY_COLUMNS:
                        new_cell.number_format = currency_format
                    elif base_col in self.TEXT_COLUMNS:
                        new_cell.number_format = '@'
                        # Forçar valor como string se já foi lido como número
                        if val is not None:
                            new_cell.value = str(val).strip()

                    if is_parent:
                        # Formatação especial para Fatura Pai
                        new_cell.font = parent_font
                        new_cell.fill = parent_fill
                    else:
                        # Para colunas adicionadas logicamente, usar a primeira coluna como modelo visual seguro.
                        col_ref = ws.cell(row=2, column=col_idx)
                        model_ref = ws.cell(row=2, column=1)
                        use_model = (base_col == CLASSIFICATION_COL) or (not col_ref.border or not col_ref.border.left.style)
                        style_source = model_ref if use_model else col_ref

                        new_cell.font = copy(style_source.font)
                        new_cell.border = copy(style_source.border)
                        new_cell.alignment = copy(style_source.alignment)
                        new_cell.protection = copy(style_source.protection)

                        if base_col not in self.CURRENCY_COLUMNS:
                            new_cell.number_format = copy(col_ref.number_format)

                    if is_separator:
                        new_cell.fill = openpyxl.styles.PatternFill(fill_type=None)
                    elif not is_parent:
                        new_cell.fill = openpyxl.styles.PatternFill(fill_type=None)

                    # Destaque de ausência (apenas na fonte, fundo permanece limpo)
                    if not is_separator:
                        is_empty = val is None or str(val).strip().lower() in ["", "nan", "nat", "none"]
                        if is_empty and base_col in {"Vencimento", "Status Pos-Faturamento"}:
                            new_cell.font = missing_font

                            uc_logical_col = ENRICHMENT_KEY if ENRICHMENT_KEY in template_col_to_idx else "CPF/CNPJ"
                            if uc_logical_col in template_col_to_idx:
                                uc_idx = template_col_to_idx[uc_logical_col]
                                uc_cell = ws.cell(row=current_row, column=uc_idx)
                                if uc_cell.comment is None:
                                    from openpyxl.comments import Comment
                                    uc_cell.comment = Comment(
                                        "⚠ Dado ausente na Gestão de Cobrança para este período",
                                        "Sistema MC"
                                    )
                                    uc_cell.comment.width = 200
                                    uc_cell.comment.height = 50

                current_row += 1
                total_rows_written += 1

        # Remover aba temporária
        wb.remove(template_ws)

        # Resumo Executivo
        if incluir_resumo:
            resumo_ws = wb.create_sheet("Resumo Executivo")
            resumo_ws.append(["Resumo Executivo"])
            resumo_ws.append([])

            is_parent = data_to_insert.get(PARENT_ROW_FLAG, pd.Series(False, index=data_to_insert.index)).astype(bool)
            is_separator = data_to_insert.get(SEPARATOR_ROW_FLAG, pd.Series(False, index=data_to_insert.index)).astype(bool)
            raw_rows = data_to_insert[~(is_parent | is_separator)]

            def _parse_num(v):
                if pd.isna(v): return 0.0
                if isinstance(v, (int, float)): return float(v)
                s = str(v).strip()
                if s in ["", "-", "--"]: return 0.0
                if "," in s: s = s.replace(".", "").replace(",", ".")
                try: return float(s)
                except: return 0.0

            eco_total = raw_rows["Ganho total Padrão"].apply(_parse_num).sum() if "Ganho total Padrão" in raw_rows.columns else 0.0
            # Total de faturamento deve usar valor monetário da fatura, não tarifa unitária.
            fat_total = raw_rows["Valor Enviado Emissão"].apply(_parse_num).sum() if "Valor Enviado Emissão" in raw_rows.columns else 0.0

            resumo_ws.append(["Soma da Economia Gerada (R$):", eco_total])
            resumo_ws.append(["Soma da Fatura Raízen (R$):", fat_total])
            resumo_ws.append([])
            resumo_ws.append(["Situação do Pagamento", "Contagem"])
            
            if "Status Pos-Faturamento" in raw_rows.columns:
                counts = raw_rows["Status Pos-Faturamento"].value_counts()
                for status, count in counts.items():
                    resumo_ws.append([status, count])

            # Formatação básica do resumo
            for row_idx in [1, 4, 5, 7]:
                resumo_ws.cell(row=row_idx, column=1).font = openpyxl.styles.Font(bold=True)
            resumo_ws.cell(row=4, column=2).number_format = currency_format
            resumo_ws.cell(row=5, column=2).number_format = currency_format

        logger.info("Planilha gerada com %d linhas de dados em %d separadores.", total_rows_written, len(df_groups))

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
