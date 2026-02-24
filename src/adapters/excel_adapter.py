import pandas as pd
from typing import List, Dict, Any

class BaseExcelReader:
    """Adaptador para leitura da planilha gd_gestao_cobranca."""
    
    def __init__(self, file_path_or_buffer: Any):
        """Inicializa o leitor e já carrega o DataFrame. Suporta string(path) ou UploadedFile."""
        self.df = pd.read_excel(file_path_or_buffer)
        
    def get_clients(self) -> List[str]:
        """Retorna lista de clientes (Nome) únicos na base, ordenados."""
        if 'Nome' not in self.df.columns:
            return []
        clients = self.df['Nome'].dropna().unique().tolist()
        return sorted([str(c) for c in clients])

    def get_periods(self) -> List[str]:
        """Retorna lista de períodos (Mês de Referência) únicos."""
        col = 'Mês de Referência'
        if col not in self.df.columns:
            return []
        periods = self.df[col].dropna().unique().tolist()
        return sorted([str(p) for p in periods])

    def filter_data(self, clients: List[str], period: str) -> pd.DataFrame:
        """Filtra o DataFrame pelos clientes e período especificados."""
        mask = pd.Series(True, index=self.df.index)
        
        if clients:
            mask = mask & (self.df['Nome'].isin(clients))
            
        if period:
            mask = mask & (self.df['Mês de Referência'] == period)
            
        return self.df[mask].copy()

class TemplateExcelWriter:
    """Adaptador para escrever dados no template mc.xlsx."""
    
    def __init__(self, template_path_or_buffer: Any):
        self.template_source = template_path_or_buffer
        
    def generate_bytes(self, data_to_insert: pd.DataFrame, column_mapping: Dict[str, str]) -> bytes:
        """
        Lê o template, insere as linhas filtradas e retorna os bytes do Excel gerado.
        Ideal para Streamlit não precisar salvar em disco.
        """
        import io
        import openpyxl
        
        # Load the template workbook
        wb = openpyxl.load_workbook(self.template_source)
        
        # Assume writing to the first active sheet
        ws = wb.active
        
        # Find the header row in template
        # Usually it's row 1, but we should find the column indexes
        header_row_idx = 1
        template_headers = {cell.value: idx for idx, cell in enumerate(ws[header_row_idx], 1) if cell.value}
        
        # Determine the next empty row
        start_row = ws.max_row + 1
        
        # Se for o template original com openpyxl e ele tiver muito pouca coisa, max_row costuma ser 2
        # É seguro escrever iterando pelas rows geradas.
        if start_row <= 2 and ws.cell(row=2, column=1).value is None:
            # Overwrite example rows if any
            start_row = 2
            
        # Para cada linha do dataframe
        for idx, row in data_to_insert.iterrows():
            current_row = start_row + idx if isinstance(idx, int) else start_row
            # Reset idx for writing (it might be the dataframe original index)
            # So we use an explicit counter
            pass
            
        current_row = start_row
        from copy import copy
        
        for _, row in data_to_insert.iterrows():
            for base_col, template_col in column_mapping.items():
                if base_col in row and template_col in template_headers:
                    col_idx = template_headers[template_col]
                    val = row[base_col]
                    # Handle NaNs
                    if pd.isna(val):
                        val = None
                    
                    new_cell = ws.cell(row=current_row, column=col_idx, value=val)
                    
                    # Copiar a formatação da linha 2 (referência do template) para as novas linhas
                    if current_row > 2:
                        ref_cell = ws.cell(row=2, column=col_idx)
                        if ref_cell.has_style:
                            new_cell.font = copy(ref_cell.font)
                            new_cell.border = copy(ref_cell.border)
                            new_cell.fill = copy(ref_cell.fill)
                            new_cell.number_format = copy(ref_cell.number_format)
                            new_cell.protection = copy(ref_cell.protection)
                            new_cell.alignment = copy(ref_cell.alignment)

            current_row += 1
            
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
