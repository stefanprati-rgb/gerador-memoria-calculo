"""
Inspeção leve da coluna Excecao Fat. usando openpyxl (sem carregar tudo).
"""
import openpyxl

FILE = r"C:\Projetos\memoria_de_calculo\Balanco_Energetico_Raizen.xlsm"
SHEET = "Balanco Operacional"

wb = openpyxl.load_workbook(FILE, read_only=True, data_only=True)
ws = wb[SHEET]

# 1. Detectar header — procurar "No. UC" nas primeiras 20 linhas
header_row = None
header_map = {}
for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), 1):
    vals = [str(c.value).strip() if c.value else "" for c in row]
    if "No. UC" in vals and "CPF/CNPJ" in vals:
        header_row = row_idx
        header_map = {v: i for i, v in enumerate(vals) if v}
        break

print(f"Header na linha: {header_row}")

# 2. Pegar índice de Excecao Fat.
exc_idx = header_map.get("Excecao Fat.")
print(f"Índice de 'Excecao Fat.': {exc_idx}")

# 3. Contar valores dessa coluna
from collections import Counter
counter = Counter()
total = 0
for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
    if exc_idx is not None and exc_idx < len(row):
        val = row[exc_idx]
        counter[str(val) if val else "NaN/Vazio"] += 1
    total += 1

print(f"\nTotal de registros: {total}")
print("\n=== Valores de 'Excecao Fat.' ===")
for k, v in counter.most_common():
    print(f"  {k}: {v}")

wb.close()
