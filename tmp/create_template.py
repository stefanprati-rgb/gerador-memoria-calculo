import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def create_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Memoria de Calculo"

    headers = [
        "Referencia", "No. UC", "CPF/CNPJ", "Razao Social", "Distribuidora",
        "Cred. Consumido Raizen", "Desconto Contratado",
        "Vencimento", "Status Pos-Faturamento", "Boleto Raizen",
        "Tarifa Raizen", "Custo c/ GD", "Custo s/ GD",
        "Ganho total Padrão"
    ]

    # Style definitions (matching expected look)
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    border_thin = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    alignment_center = Alignment(horizontal='center', vertical='center')

    # Row 1: Headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border_thin
        cell.alignment = alignment_center
        
        # Adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 5

    # Row 2: Reference Style Row
    # TemplateExcelWriter.generate_bytes (lines 366-376) copies style from Row 2
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.border = border_thin
        
        # Numeric columns alignment
        if headers[col_idx-1] in ["Cred. Consumido Raizen", "Boleto Raizen", "Tarifa Raizen", "Custo c/ GD", "Custo s/ GD", "Ganho total Padrão"]:
             cell.alignment = Alignment(horizontal='right')
        else:
             cell.alignment = Alignment(horizontal='left')

    template_path = "mc.xlsx"
    wb.save(template_path)
    print(f"Template '{template_path}' created successfully.")

if __name__ == "__main__":
    create_template()
