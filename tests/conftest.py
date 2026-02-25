"""
Fixtures compartilhadas para os testes do projeto Memória de Cálculo.
Adaptadas para a nova base Balanço Energético com suporte a agrupamento.
"""

import pytest
import pandas as pd
import openpyxl
import io


@pytest.fixture
def sample_base_df():
    """DataFrame simulando a planilha Balanço Energético com dados mínimos.
    Inclui linhas com e sem agrupamento para testar ambos os cenários.
    """
    return pd.DataFrame({
        "Referencia": ["2026-01-01", "2026-01-01", "2026-02-01", "2026-01-01", "2026-01-01", "2026-01-01"],
        "No. UC": ["UC001", "UC002", "UC003", "UC004", "UC005", "UC006"],
        "CPF/CNPJ": ["11111111000101", "22222222000102", "11111111000101", "33333333000103", "33333333000103", "44444444000104"],
        "Razao Social": ["Cliente Alpha", "Cliente Beta", "Cliente Alpha", "Cliente Gamma", "Cliente Gamma", "Cliente Delta"],
        "Distribuidora": ["CEMIG", "CPFL", "CEMIG", "CEMIG", "CEMIG", "LIGHT"],
        "Cred. Consumido Raizen": [1500.5, 2300.0, 1600.75, 800.0, 700.0, 500.0],
        "Desconto Contratado": ["15%", "10%", "15%", "20%", "20%", "12%"],
        "Vencimento": ["2026-01-15", "2026-01-20", "2026-02-15", "2026-01-10", "2026-01-10", "2026-01-25"],
        "Status Pos-Faturamento": ["Pago", "Pendente", "Pendente", "Pago", "Pago", "Pago"],
        "Boleto Raizen": [350.00, 520.50, 370.25, 200.00, 180.00, 150.00],
        "Tarifa Raizen": [0.85, 0.92, 0.85, 0.78, 0.78, 0.80],
        "Custo c/ GD": [1275.43, 2116.00, 1360.64, 624.00, 546.00, 400.00],
        "Custo s/ GD": [1625.43, 2636.50, 1730.89, 824.00, 726.00, 550.00],
        "Ganho total Padrão": [350.00, 520.50, 370.25, 200.00, 180.00, 150.00],
        # Gamma (2 UCs) e Delta (1 UC sozinha) marcados como Agrupamento
        "Excecao Fat.": [None, None, None, "Agrupamento", "Agrupamento", "Agrupamento"],
    })


@pytest.fixture
def sample_base_xlsx(sample_base_df, tmp_path):
    """Arquivo .xlsx temporário simulando a aba Balanço Operacional com header na linha 6.
    As primeiras 5 linhas contêm dados auxiliares, o header real está na linha 6 (index 5).
    """
    path = tmp_path / "base_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balanco Operacional"

    # Linhas 1-5: dados auxiliares (simulando o que existe na base real)
    ws.cell(row=1, column=1, value="Informação auxiliar 1")
    ws.cell(row=2, column=1, value="Informação auxiliar 2")
    ws.cell(row=3, column=1, value="Informação auxiliar 3")
    ws.cell(row=4, column=1, value="Informação auxiliar 4")
    ws.cell(row=5, column=1, value="Informação auxiliar 5")

    # Linha 6: headers
    for col_idx, col_name in enumerate(sample_base_df.columns, 1):
        ws.cell(row=6, column=col_idx, value=col_name)

    # Linhas 7+: dados
    for row_idx, (_, row) in enumerate(sample_base_df.iterrows(), 7):
        for col_idx, col_name in enumerate(sample_base_df.columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])

    wb.save(path)
    return str(path)


@pytest.fixture
def sample_template_xlsx(tmp_path):
    """Arquivo template mc.xlsx mínimo para testes."""
    path = tmp_path / "mc_test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active

    # Headers do template (com espaços extras como no real para testar normalização)
    headers = [
        "Data  Ref", "UC", "CNPJ", "Razão Social", "Distribuidora",
        "Energia compensada pela Raízen (kWh)", "Regra aplicada",
        "Vencimento ", "Status financeiro", "Boleto faturado (R$)",
        "Tafira Distribuidora", "Custo com GD R$", "Custo sem GD R$",
        "Economia (R$)"
    ]

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    wb.save(path)
    return str(path)
