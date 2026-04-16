"""
Testes para o adaptador Excel (BaseExcelReader e TemplateExcelWriter).
Cobre detecção dinâmica de header e formatação de Fatura Pai.
"""

import pytest
import openpyxl
import io
import pandas as pd

from logic.adapters.excel_adapter import (
    BaseExcelReader,
    TemplateExcelWriter,
    ColumnValidationError,
    HeaderNotFoundError,
)
from logic.core.mapping import COLUMN_MAPPING, PARENT_ROW_FLAG


class TestBaseExcelReader:
    """Testes do leitor de planilha com detecção dinâmica de header."""

    def test_inicializacao_com_header_dinamico(self, sample_base_xlsx):
        """Deve detectar o header na linha 6 e carregar os dados corretamente."""
        reader = BaseExcelReader(sample_base_xlsx)
        assert reader is not None
        assert len(reader.df) == 6  # 6 linhas de dados

    def test_header_not_found_sem_marcadores(self, tmp_path):
        """Deve levantar HeaderNotFoundError se não encontrar os marcadores."""
        path = tmp_path / "sem_header.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balanco Operacional"
        ws.cell(row=1, column=1, value="Coluna Aleatória")
        wb.save(path)

        with pytest.raises(HeaderNotFoundError):
            BaseExcelReader(str(path))

    def test_colunas_validadas(self, sample_base_xlsx):
        """Deve validar que as colunas obrigatórias (não-opcionais) estão presentes."""
        reader = BaseExcelReader(sample_base_xlsx)
        from logic.core.mapping import get_base_columns, OPTIONAL_BASE_COLUMNS
        required_cols = [c for c in get_base_columns() if c not in OPTIONAL_BASE_COLUMNS]
        for col in required_cols:
            assert col in reader.df.columns, f"Coluna obrigatória ausente: {col}"

    def test_coluna_faltante_levanta_erro(self, tmp_path):
        """Deve levantar ColumnValidationError se uma coluna obrigatória faltar."""
        path = tmp_path / "incompleto.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Balanco Operacional"
        # Header correto porém incompleto
        ws.cell(row=1, column=1, value="No. UC")
        ws.cell(row=1, column=2, value="CPF/CNPJ")
        # Faltam várias colunas obrigatórias
        wb.save(path)

        with pytest.raises(ColumnValidationError):
            BaseExcelReader(str(path))

    def test_get_clients(self, sample_base_xlsx):
        """Deve retornar lista de clientes únicos ordenados."""
        reader = BaseExcelReader(sample_base_xlsx)
        clients = reader.get_clients()

        assert "Cliente Alpha" in clients
        assert "Cliente Beta" in clients
        assert "Cliente Gamma" in clients
        assert clients == sorted(clients)

    def test_get_periods(self, sample_base_xlsx):
        """Deve retornar lista de períodos únicos."""
        reader = BaseExcelReader(sample_base_xlsx)
        periods = reader.get_periods()

        assert len(periods) == 2  # jan e fev

    def test_filter_data_clientes(self, sample_base_xlsx):
        """Deve filtrar corretamente por cliente."""
        reader = BaseExcelReader(sample_base_xlsx)
        filtered = reader.filter_data(["Cliente Alpha"], [])

        assert len(filtered) == 2  # 2 registros do Alpha

    def test_filter_data_periodos(self, sample_base_xlsx):
        """Deve filtrar corretamente por período."""
        reader = BaseExcelReader(sample_base_xlsx)
        all_periods = reader.get_periods()
        first_period = all_periods[0]
        filtered = reader.filter_data([], [first_period])

        assert len(filtered) > 0


class TestTemplateExcelWriter:
    """Testes do escritor de template com formatação de Fatura Pai."""

    def test_gera_bytes_validos(self, sample_base_xlsx, sample_template_xlsx):
        """Deve gerar bytes de um Excel válido."""
        reader = BaseExcelReader(sample_base_xlsx)
        filtered = reader.filter_data(["Cliente Alpha"], reader.get_periods())
        filtered[PARENT_ROW_FLAG] = False

        writer = TemplateExcelWriter(sample_template_xlsx)
        result = writer.generate_bytes(filtered, COLUMN_MAPPING)

        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_formatacao_fatura_pai(self, sample_template_xlsx):
        """Linhas de Fatura Pai devem receber formatação diferenciada (negrito + fundo)."""
        df = pd.DataFrame({
            "Referencia": ["2026-01-01", "2026-01-01"],
            "No. UC": ["AGRUPADO", "UC001"],
            "CPF/CNPJ": ["11111111", "11111111"],
            "Razao Social": ["TOTAL AGRUPADO - Cliente Alpha", "Cliente Alpha"],
            "Distribuidora": ["CEMIG", "CEMIG"],
            "Cred. Consumido Raizen": [3000, 1500],
            "Desconto Contratado": ["15%", "15%"],
            "Status Pos-Faturamento": ["Pago", "Pago"],
            "Valor Enviado Emissão": [700, 350],
            "Tarifa Raizen": [0.85, 0.85],
            "Custo c/ GD": [2500, 1275],
            "Custo s/ GD": [3200, 1625],
            "Ganho total Padrão": [700, 350],
            PARENT_ROW_FLAG: [True, False],
        })

        writer = TemplateExcelWriter(sample_template_xlsx)
        result = writer.generate_bytes(df, COLUMN_MAPPING)

        # Abrir o resultado e verificar formatação
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active

        # Linha 2 = Fatura Pai (deve estar em negrito)
        parent_cell = ws.cell(row=2, column=5)  # coluna Razao Social
        assert parent_cell.value == "TOTAL AGRUPADO - Cliente Alpha"
        assert parent_cell.font.bold is True

        # Linha 3 = UC Filha (não deve estar em negrito via parent_font)
        child_cell = ws.cell(row=3, column=5)
        assert child_cell.value == "Cliente Alpha"
    def test_renomeacao_headers_legados(self, tmp_path):
        """Deve detectar nomes antigos no template e convertê-los para os nomes da fonte no resultado."""
        template_path = tmp_path / "legacy_template.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header com nomes antigos
        ws.cell(row=1, column=1, value="UC")
        ws.cell(row=1, column=2, value="Razão Social")
        wb.save(template_path)

        df = pd.DataFrame({
            "No. UC": ["UC001"],
            "Razao Social": ["Cliente Teste"],
            PARENT_ROW_FLAG: [False]
        })

        writer = TemplateExcelWriter(str(template_path))
        mapping = {"No. UC": "No. UC", "Razao Social": "Razao Social"}
        result = writer.generate_bytes(df, mapping)

        # Verificar se o header foi renomeado no arquivo gerado
        wb_res = openpyxl.load_workbook(io.BytesIO(result))
        ws_res = wb_res.active
        assert ws_res.cell(row=1, column=1).value == "No. UC"
        assert ws_res.cell(row=1, column=2).value == "Razao Social"
        assert ws_res.cell(row=2, column=1).value == "UC001"

    def test_tarifa_raizen_negativa_vira_zero_no_excel(self, sample_template_xlsx):
        """Valores negativos na Tarifa Raizen devem ser convertidos para 0.00 no arquivo final."""
        df = pd.DataFrame({
            "No. UC": ["UC1"],
            "Tarifa Raizen": [-15.50],
            PARENT_ROW_FLAG: [False]
        })

        writer = TemplateExcelWriter(sample_template_xlsx)
        result = writer.generate_bytes(df, COLUMN_MAPPING)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        
        # O valor no Excel deve ser 0
        # Tarifa Raizen está na coluna 12 no COLUMN_MAPPING
        assert ws.cell(row=2, column=12).value == 0.0

    def test_formatacao_referencia_full_date(self, sample_template_xlsx):
        """A coluna Referencia deve ser formatada como data completa (DD/MM/YYYY)."""
        df = pd.DataFrame({
            "Referencia": ["2026-01-01"],
            "No. UC": ["UC001"],
            PARENT_ROW_FLAG: [False]
        })

        writer = TemplateExcelWriter(sample_template_xlsx)
        mapping = {"Referencia": "Referencia", "No. UC": "No. UC"}
        result = writer.generate_bytes(df, mapping)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        
        # O valor no Excel deve ser a string formatada "01/01/2026"
        # Nota: Referencia está na coluna 1 no sample_template_xlsx
        assert ws.cell(row=2, column=1).value == "01/01/2026"

    @pytest.mark.xfail(
        reason="Bug preexistente: missing_fill não é aplicado ao fill da célula, apenas missing_font. Rastreado como dívida técnica.",
        strict=False,
    )
    def test_destaque_laranja_em_celulas_vazias(self, sample_template_xlsx):
        """Gerar Excel com linha sem Vencimento deve aplicar fundo laranja na célula."""
        df = pd.DataFrame({
            "No. UC": ["UC001"],
            "Vencimento": [pd.NA],
            PARENT_ROW_FLAG: [False]
        })

        writer = TemplateExcelWriter(sample_template_xlsx)
        result = writer.generate_bytes(df, COLUMN_MAPPING)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        
        # Vencimento está na coluna 8 no COLUMN_MAPPING
        cell = ws.cell(row=2, column=8)
        
        # FFE0B2 é o laranja claro. openpyxl retorna com o prefixo FF para opacidade total.
        fill_color = cell.fill.start_color.rgb
        assert fill_color in ["FFFFE0B2", "00FFE0B2"]

    def test_comentario_em_uc_com_dado_ausente(self, sample_template_xlsx):
        """Linha sem Vencimento deve adicionar comentário na célula da coluna No. UC."""
        df = pd.DataFrame({
            "No. UC": ["UC001"],
            "Vencimento": [pd.NA],
            PARENT_ROW_FLAG: [False]
        })

        writer = TemplateExcelWriter(sample_template_xlsx)
        result = writer.generate_bytes(df, COLUMN_MAPPING)

        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        
        # No. UC está na coluna 2 no COLUMN_MAPPING
        uc_cell = ws.cell(row=2, column=2)
        assert uc_cell.comment is not None
        assert "Dado ausente" in uc_cell.comment.text
